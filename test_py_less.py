import re, random
from openpyxl import Workbook
from openpyxl.styles import PatternFill

test_dir = 'C:\\Users\\laser\\Desktop\\'
testfile = 'py_xls.xlsx'

book = Workbook()
ws = book.active
color = str(hex(random.randint(0, 16777215)))[2:]
print (color)
ws.sheet_properties.tabColor = color

for row in ws.iter_rows(min_row=1, max_col=5, max_row=5):
    for cell in row:
        color = str(hex(random.randint(0, 16777215)))[2:]
        cell.value = str(color)
        cell.fill = PatternFill('solid', fgColor=color)

book.save(testfile)
