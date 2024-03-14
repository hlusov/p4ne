from matplotlib import pyplot
from openpyxl import Workbook, load_workbook
import urllib.request as url

file = 'plot.xlsx'
url.urlretrieve("https://github.com/mk-99/p4ne/raw/master/Lab1.2/data_analysis_lab.xlsx", "data.xlsx")
src_xls_list = load_workbook("data.xlsx")
sheet = src_xls_list.active
def gfx(x):
    return x.value

years = list(map(gfx, sheet['A'][1:]))
temperature = list(map(gfx, sheet['C'][1:]))
activity = list(map(gfx, sheet['D'][1:]))

pyplot.plot(years, temperature, label="Относит. температура")
pyplot.plot(years, activity, label="Активность Солнца")

pyplot.xlabel('Годы')
pyplot.ylabel('Температура/Активность Солнца')
pyplot.legend(loc='upper left')

#pyplot.show()

wb = Workbook()
ws = wb.active

#for i in range(len(years)):
#    print(years[i])
#    ws['A'][i+1] = years[i]
print(ws.title)
wb.save(file)

