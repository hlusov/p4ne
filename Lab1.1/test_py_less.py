import re, random
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

#test_dir = 'C:\\Users\\laser\\Desktop\\'
testfile = 'py_xls.xlsx'

def rand():
    return "ff" + str(hex(random.randint(0, 2**24))[2:])
book = Workbook()
ws = book.active
color = rand()
print(color)
ws.sheet_properties.tabColor = color

for row in ws.iter_rows(min_row=1, max_col=5, max_row=5):
    for cell in row:
        cell.value = rand()
        print(cell.value)
        cell.fill = PatternFill('solid', fgColor=cell.value)
        ws[row, cell].font = Font(color=random.randint(2, 10))

book.save(testfile)

